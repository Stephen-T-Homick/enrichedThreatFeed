import requests
import json
import os
from datetime import datetime, timezone
from rich.console import Console
from rich.table import Table
from rich import box
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from dotenv import load_dotenv
from urllib.parse import urlparse

load_dotenv()

# API Config # 
THREATFOX_API_KEY = os.getenv("THREATFOX_API_KEY")
ABUSEIPDB_API_KEY = os.getenv("ABUSEIPDB_API_KEY")
OTX_API_KEY = os.getenv("OTX_API_KEY")
DAYS_LOOKBACK = 1
RELEVANCE_THRESHOLD = 70
#

## Excel ## 
TODAY_DATE = datetime.now(timezone.utc).strftime('%Y-%m-%d')
EXCEL_FILE = "global_ioc_hunt.xlsx"
EXCEL_SHEET = TODAY_DATE
##
RELEVANT_MALWARE = [
    "win.redline", "win.lumma", "win.agent_tesla", "win.raccoon", "win.ficker",
    "win.qbot", "win.iceid", "win.remcos", "win.cobalt_strike"
]
RELEVANT_TAGS = ["cloud", "aws", "gcp", "azure", "s3", "pgadmin", "sql", "vpn", "k8s", "stealer",
                 "linux", "unix", "mac"]
RELEVANT_TYPES = ["botnet_cc", "payload_delivery", "dropper"]

console = Console()
SEEN_IOCS = set()
KQL_IOCS = []
FQL_IOCS = []

### Feed Collectors ###
def fetch_threatfox():
    headers = {"Auth-Key": THREATFOX_API_KEY}
    payload = {"query": "get_iocs", "days": DAYS_LOOKBACK}
    r = requests.post("https://threatfox-api.abuse.ch/api/v1/", headers=headers, json=payload)
    if r.status_code == 200 and r.json().get("query_status") == "ok":
        return r.json().get("data", [])
    return []

def fetch_urlhaus():
    payload = {"urlhaus_download": "recent"}
    out = []
    try:
        r = requests.post("https://urlhaus.abuse.ch/api/", json=payload)
        if r.status_code == 200:
            json_data = r.json()
            for e in json_data.get("urls", []):
                if e["url_status"] != "online": continue
                out.append({"ioc": e["url"], "ioc_type": "url", "malware": e.get("signature", "unknown"),
                            "malware_printable": e.get("signature", "unknown"), "threat_type": "payload_delivery",
                            "threat_type_desc": "Malicious URL (URLhaus)", "tags": ["url"], "confidence_level": "high",
                            "first_seen": e.get("date_added", "")})
    except Exception as e:
        console.print(f"[red]Error fetching URLHaus feed: {e}[/red]")
    return out

def fetch_feodo():
    r = requests.get("https://feodotracker.abuse.ch/downloads/ipblocklist.csv")
    return [{"ioc": l.strip(), "ioc_type": "ip", "malware": "trickbot", "malware_printable": "TrickBot",
             "threat_type": "botnet_cc", "threat_type_desc": "Feodo Tracker C2", "tags": ["feodo"],
             "confidence_level": "high", "first_seen": TODAY_DATE} for l in r.text.splitlines() if l and not l.startswith("#")]

def fetch_tor_exit():
    r = requests.get("https://check.torproject.org/torbulkexitlist")
    return [{"ioc": ip, "ioc_type": "ip", "malware": "anonymized_traffic", "malware_printable": "Tor Exit",
             "threat_type": "anonymized_exit", "threat_type_desc": "Tor Exit Node", "tags": ["tor"],
             "confidence_level": "medium", "first_seen": TODAY_DATE} for ip in r.text.splitlines() if ip]

def fetch_alienvault_otx():
    headers = {"X-OTX-API-KEY": OTX_API_KEY}
    r = requests.get("https://otx.alienvault.com/api/v1/pulses/subscribed", headers=headers)
    out = []
    if r.status_code == 200:
        for pulse in r.json().get("results", [])[:10]:
            for ind in pulse.get("indicators", []):
                out.append({"ioc": ind["indicator"], "ioc_type": ind["type"].lower(),
                            "malware": pulse.get("name", "unknown"),
                            "malware_printable": pulse.get("name", "unknown"),
                            "threat_type": "otx", "threat_type_desc": pulse.get("description", ""),
                            "tags": pulse.get("tags", []), "confidence_level": "medium",
                            "first_seen": pulse.get("created", "")})
    return out

def fetch_spamhaus_drop():
    r = requests.get("https://www.spamhaus.org/drop/drop.txt")
    return [{"ioc": l.split(";")[0], "ioc_type": "ip", "malware": "spamhaus", "malware_printable": "DROP List",
             "threat_type": "criminal_infra", "threat_type_desc": "Known bad IP", "tags": ["spamhaus"],
             "confidence_level": "high", "first_seen": TODAY_DATE} for l in r.text.splitlines() if not l.startswith(";") and l.strip()]
###

#### Utility ####
def defang(ioc, ioc_type):
    if ioc_type.startswith("ip"):
        return ioc.split(":")[0].replace(".", "[.]")
    if ioc_type in ["domain", "url"]:
        return ioc.replace(".", "[.]")
    return ioc

def relevance(i):
    s = 30
    if i.get("malware", "").lower() in RELEVANT_MALWARE: s += 40
    if any(t in RELEVANT_TAGS for t in (i.get("tags") or [])): s += 20
    if i.get("threat_type", "").lower() in RELEVANT_TYPES: s += 10
    return min(s, 100)

def abuseipdb(ioc):
    try:
        r = requests.get("https://api.abuseipdb.com/api/v2/check",
                         headers={"Key": ABUSEIPDB_API_KEY, "Accept": "application/json"},
                         params={"ipAddress": ioc, "maxAgeInDays": 30})
        if r.status_code == 200:
            return 'y' if r.json().get("data", {}).get("abuseConfidenceScore", 0) >= 25 else 'n'
    except:
        pass
    return 'n'

def print_ioc(i, score):
    table = Table(title=f"Relevant IOC (Score: {score})", box=box.SIMPLE_HEAVY)
    table.add_column("Field", style="bold green")
    table.add_column("Value", style="white")
    table.add_row("IOC", i["ioc"])
    table.add_row("IOC Type", i.get("ioc_type", ""))
    table.add_row("Malware", i.get("malware_printable", ""))
    table.add_row("Threat Type", i.get("threat_type_desc", ""))
    table.add_row("Confidence", str(i.get("confidence_level", "")))
    table.add_row("First Seen", i.get("first_seen", ""))
    tags = i.get("tags") or []
    table.add_row("Tags", ", ".join(tags if isinstance(tags, list) else [str(tags)]))
    console.print(table)

def build_hunting_queries():
    if not KQL_IOCS and not FQL_IOCS:
        return
    kql = " OR ".join([
        f"destination.domain: \"{ioc}\"" if t == "domain" else
        f"url.full: \"{ioc}\"" if t == "url" else
        f"destination.ip: \"{ioc.split(':')[0]}\"" for ioc, t in KQL_IOCS
    ])
    fql = " OR ".join([
        f"DomainName={ioc}" if t == "domain" else
        f"RemoteIP={ioc.split(':')[0]}" if t == "ip:port" else
        f"DomainName={urlparse(ioc).hostname}" if t == "url" else "" for ioc, t in FQL_IOCS
    ])
    console.rule("[bold red]Detection Hunting Queries")
    console.print("[bold cyan]Elastic KQL:[/bold cyan]")
    console.print(kql)
    console.print("\n[bold cyan]CrowdStrike FQL:[/bold cyan]")
    console.print(fql)

def write_excel(i, score, abuse):
    headers = ["IOC", "IOC Type", "Malware", "Threat Type", "Confidence", "First Seen", "Tags", "Score", "AbuseIPDB Hit"]
    v = defang(i["ioc"], i.get("ioc_type"))
    row = [v, i.get("ioc_type", ""), i.get("malware_printable", ""), i.get("threat_type_desc", ""),
           i.get("confidence_level", ""), i.get("first_seen", ""), ", ".join(i.get("tags", []) or []), score, abuse]
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = Workbook(); wb.remove(wb.active)
    if EXCEL_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(EXCEL_SHEET); ws.append(headers)
        for c in range(1, len(headers)+1): ws[f"{get_column_letter(c)}1"].font = Font(bold=True)
    else:
        ws = wb[EXCEL_SHEET]
    ws.append(row); wb.save(EXCEL_FILE)

def main():
    feeds = fetch_threatfox() + fetch_urlhaus() + fetch_feodo() + fetch_tor_exit() + fetch_alienvault_otx() + fetch_spamhaus_drop()
    console.print(f"[bold cyan]Fetched {len(feeds)} IOCs from all sources[/bold cyan]")
    hits = 0
    for i in feeds:
        k = (i["ioc"], i.get("ioc_type"))
        if k in SEEN_IOCS: continue
        SEEN_IOCS.add(k)
        s = relevance(i)
        if s < RELEVANCE_THRESHOLD: continue
        abuse = 'n'
        if i.get("ioc_type") in ["ip", "ip:port"]:
            abuse = abuseipdb(i["ioc"].split(":")[0])
        print_ioc(i, s)
        write_excel(i, s, abuse)
        if i.get("ioc_type") in ["domain", "url", "ip:port", "ip"]:
            KQL_IOCS.append((i["ioc"], i.get("ioc_type")))
            FQL_IOCS.append((i["ioc"], i.get("ioc_type")))
        hits += 1
    if hits:
        build_hunting_queries()
        console.print(f"[green]Wrote {hits} global IOCs to {EXCEL_FILE}[/green]")
    else:
        console.print("[yellow]No relevant IOCs scored above threshold.[/yellow]")

if __name__ == '__main__':
    main()
####
